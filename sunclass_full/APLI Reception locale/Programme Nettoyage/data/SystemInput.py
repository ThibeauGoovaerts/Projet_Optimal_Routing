import os
import openpyxl
import pyexcel as p
import time
import shutil
from retrieveData import readParams, getAddedReloues, getAddDeparts, getAddCheckCaution, getNotCleaned, getNeedControle
from export import make_winter_plan
from tkinter import Tk, HORIZONTAL, Button, Label, ttk, Text, Entry, IntVar, Toplevel


def makeDirectory(file_name, desktop):
	# If the directory already exists, we need to remove it, this sometimes bugs, so we do a while loop
	while os.path.isdir(desktop + f"/schoonmaaklijst/{file_name}"):
		shutil.rmtree(desktop + f"/schoonmaaklijst/{file_name}")
	
	# We create the new directory, this is also in a while statement, sometimes this also bugs, this has to do
	# with the transmission and cooperation between os and python
	while not os.path.isdir(desktop + f"/schoonmaaklijst/{file_name}"):
		os.mkdir(desktop + f"/schoonmaaklijst/{file_name}")
	return


def moveExcel(file_name, desktop):
	# We take the excel sheet and move it into our newly created directory
	os.rename(desktop + "/schoonmaaklijst.xlsx", desktop + f"/schoonmaaklijst/{file_name}/schoonmaaklijst.xlsx")
	# We make a parameter my_path which will be the path to the excel file
	my_path = desktop + f"/schoonmaaklijst/{file_name}/schoonmaaklijst.xlsx"
	return my_path


def checkExistenceExcel(desktop):
	# We check if the file "schoonmaaklijst.xls" as the export is called is present on the desktop
	# Remark that this is the old extension of excel
	if os.path.isfile(desktop + "/schoonmaaklijst.xls"):
		# If it is present we change it to the recent version of excel, one we can use on the computer and in the program
		p.save_book_as(file_name=desktop + "/schoonmaaklijst.xls", dest_file_name=desktop + "/schoonmaaklijst.xlsx")
		# We remove the old version of the excel file
		os.remove(desktop + "/schoonmaaklijst.xls")
	
	# If the main map does not exist, we create it
	if not os.path.isdir(desktop + "/schoonmaaklijst"):
		os.mkdir(desktop + "/schoonmaaklijst")
	
	# We check if a excel file exist in the main map, if it does, we delete it since we won't use it and it will
	# create confusion
	if os.path.isfile(desktop + "/schoonmaaklijst/schoonmaaklijst.xlsx"):
		os.remove(desktop + "/schoonmaaklijst/schoonmaaklijst.xlsx")
	if not os.path.isfile(desktop + "/schoonmaaklijst.xlsx"):
		return False
	return True


def challenge_acceuils(acceuils):
	not_renting = readParams(bungalows_not_renting=True)
	a = [j for j in acceuils if j in not_renting]
	for k in a:
		acceuils.remove(k)
	return acceuils


def readBungalows(my_path, desktop):
	# We re-open up the excel sheet
	wb = openpyxl.load_workbook(my_path)
	sht = wb.active
	date = sht.cell(7, 4).value
	file_name = str(date)
	max_rows = sht.max_row
	
	# Initialisation of the array with the bungalows
	index_of_bungalows = []
	
	# We append all bungalows in this row
	for i in range(1, max_rows):
		v = str(sht.cell(i, 2).value)
		if 'BNG' in v:
			index_of_bungalows.append(i)
	# We initialise the three different types of bungalows
	acceuils = []
	reloue = []
	departs = []
	
	# We initialise all the departures and all the arrivals for the check-ups
	check_Caution = []
	
	# We initialise the different types of items we can have to deal with
	ka_c = []
	ku_c = []
	ks_c = []
	kb_c = []
	ko_c = []
	ka_m = []
	ku_m = []
	ks_m = []
	kb_m = []
	ko_m = []
	
	boekingsnummers_bungnummer = []
	
	winter_aankomst = []
	winter_vertrek = []
	
	# We now go and check out every individual bungalow in the dataset
	for index in index_of_bungalows:
		# We extract the bungalows number
		nr = str(sht.cell(index, 2).value)
		nr = int(nr[3:])
		# We check the three parameters aankomst, vertrek and schoonmaak
		aankomst = str(sht.cell(index, 3).value)
		vertrek = str(sht.cell(index, 4).value)
		schoonmaak = str(sht.cell(index, 5).value)
		boekingsnummer = str(sht.cell(index, 1).value)
		# We put those three parameters into a boolean, the params will be "True" if the value if "Ja",
		# else it will be "Nee" or empty, but that doesn't matter
		aa = aankomst == 'Ja'
		vv = vertrek == 'Ja'
		ss = schoonmaak == 'Ja'
		
		if aa:
			winter_aankomst.append(nr)
		if vv:
			winter_vertrek.append(nr)
		
		# We differentiate the three different types with the booleans
		file_name_stretched = file_name.split("-")
		if aa and vv and ss:
			reloue.append(nr)
			boekingsnummers_bungnummer.append({
				'bung_nummer': nr,
				'boekingsnummer': boekingsnummer
			})
		elif vv and ss:
			departs.append(nr)
			boekingsnummers_bungnummer.append({
				'bung_nummer': nr,
				'boekingsnummer': boekingsnummer
			})
		elif aa and not ss:
			acceuils.append(nr)
		elif aa and ss:
			reloue.append(nr)
			boekingsnummers_bungnummer.append({
				'bung_nummer': nr,
				'boekingsnummer': boekingsnummer
			})
		if vv and ss:
			check_Caution.append(nr)
		
		# We look at all the possible items in the bungalow and append them if needed to the necessary list
		if int(sht.cell(index, 11).value) != 0:
			ka_m.append({'b': nr, 'nombre': int(sht.cell(index, 11).value)})
		if int(sht.cell(index, 12).value) != 0:
			ku_m.append({'b': nr, 'nombre': int(sht.cell(index, 12).value)})
		if int(sht.cell(index, 13).value) != 0:
			ks_m.append({'b': nr, 'nombre': int(sht.cell(index, 13).value)})
		if int(sht.cell(index, 14).value) != 0:
			kb_m.append({'b': nr, 'nombre': int(sht.cell(index, 14).value)})
		if int(sht.cell(index, 15).value) != 0:
			ko_m.append({'b': nr, 'nombre': int(sht.cell(index, 15).value)})
		if int(sht.cell(index, 21).value) != 0:
			ka_c.append({'b': nr, 'nombre': int(sht.cell(index, 21).value)})
		if int(sht.cell(index, 22).value) != 0:
			ku_c.append({'b': nr, 'nombre': int(sht.cell(index, 22).value)})
		if int(sht.cell(index, 23).value) != 0:
			ks_c.append({'b': nr, 'nombre': int(sht.cell(index, 23).value)})
		if int(sht.cell(index, 24).value) != 0:
			kb_c.append({'b': nr, 'nombre': int(sht.cell(index, 24).value)})
		if int(sht.cell(index, 25).value) != 0:
			ko_c.append({'b': nr, 'nombre': int(sht.cell(index, 25).value)})
	
	make_winter_plan(winter_aankomst, winter_vertrek, file_name, file_name)
	# We create a file where we will output the different items in a nice format
	file = open(desktop + f"/schoonmaaklijst/{file_name}/TD.txt", "w")
	file.write("Aller Chercher: \n")
	file.write("-----------------\n")
	
	# We write all the items (if there are any) in the file
	if len(ka_c) != 0:
		file.write("Bain enfant: ")
		file.write("\n")
		for i in ka_c:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ku_c) != 0:
		file.write("Buggy: ")
		file.write("\n")
		for i in ku_c:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ks_c) != 0:
		file.write("Chaise enfant: ")
		file.write("\n")
		for i in ks_c:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(kb_c) != 0:
		file.write("Lit Enfant: ")
		file.write("\n")
		for i in kb_c:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ko_c) != 0:
		file.write("Box Enfant: ")
		file.write("\n")
		for i in ko_c:
			file.write("- " + str(i))
			file.write("\n")
	
	file.write("\n")
	file.write("\n")
	file.write("\n")
	file.write("Aller Mettre: \n")
	file.write("-----------------\n")
	file.write("\n")
	if len(ka_m) != 0:
		file.write("Bain enfant: ")
		file.write("\n")
		for i in ka_m:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ku_m) != 0:
		file.write("Buggy: ")
		file.write("\n")
		for i in ku_m:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ks_m) != 0:
		file.write("Chaise enfant: ")
		file.write("\n")
		for i in ks_m:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(kb_m) != 0:
		file.write("Lit Enfant: ")
		file.write("\n")
		for i in kb_m:
			file.write("- " + str(i))
			file.write("\n")
	
	if len(ko_m) != 0:
		file.write("Box Enfant: ")
		file.write("\n")
		for i in ko_m:
			file.write("- " + str(i))
			file.write("\n")
	
	# We close the file
	file.close()
	
	added_reloues = getAddedReloues()
	
	for b in added_reloues:
		if b not in reloue:
			reloue.append(b)
			if b in departs:
				departs.remove(b)
			if b in acceuils:
				acceuils.remove(b)
	
	added_departs = getAddDeparts()
	
	for b in added_departs:
		if b not in departs:
			departs.append(b)
			if b in reloue:
				departs.remove(b)
			if b in acceuils:
				acceuils.remove(b)
	
	check_Caution += getAddCheckCaution()
	
	not_cleaned = getNotCleaned()
	
	for bungalow in not_cleaned:
		if bungalow in acceuils:
			reloue.append(bungalow)
			acceuils.remove(bungalow)
		elif bungalow in reloue:
			i = 1
		else:
			departs.append(bungalow)
	
	needControle = getNeedControle()
	
	for b in needControle:
		if b in acceuils:
			needControle.remove(b)
		elif b in reloue:
			acceuils.append(b)
			reloue.remove(b)
			needControle.remove(b)
		elif b in departs:
			departs.remove(b)
	
	acceuils = challenge_acceuils(acceuils)
	
	reloue = list(dict.fromkeys(reloue))
	departs = list(dict.fromkeys(departs))
	needControle = list(dict.fromkeys(needControle))
	
	for i in reloue:
		if i not in check_Caution and i not in not_cleaned:
			check_Caution.append(i)
	for i in departs:
		if i not in check_Caution and i not in not_cleaned:
			check_Caution.append(i)
	
	check_Caution = list(dict.fromkeys(check_Caution))
	return reloue, departs, acceuils, check_Caution, boekingsnummers_bungnummer, needControle


def readExcel():
	desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
	
	exists = checkExistenceExcel(desktop)
	
	if not exists:
		print(
			'>> Het EXCEL bestand is niet goed gedownload vanuit stratech, Ik kan het niet vinden, gelieve opnieuw te proberen!\n')
		time.sleep(3)
		return None, None, None, None, None, None, None
	elif exists:
		# We open the excel worksheet
		wb = openpyxl.load_workbook(desktop + "/schoonmaaklijst.xlsx")
		sht = wb.active
		# We find the date and make that into the file_name, we close the wb directly after
		date = sht.cell(7, 4).value
		file_name = str(date)
		wb.close()
		makeDirectory(file_name, desktop)
		path_to_excel = moveExcel(file_name, desktop)
		reloue, departs, acceuils, check_Caution, boekingsnummers_bungnummer, needControle = readBungalows(
			path_to_excel, desktop)
		return reloue, departs, acceuils, check_Caution, boekingsnummers_bungnummer, file_name, needControle
